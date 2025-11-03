from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from typing import Optional
from datetime import date, datetime
import io

# Excel libraries
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF libraries
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from app.models.order import Order
from app.models.service import Service


def format_location(location) -> str:
    """Format a Location object as 'City, State, Country' or 'City, Country'."""
    if not location:
        return ""
    parts = [location.city]
    if location.state:
        parts.append(location.state)
    parts.append(location.country)
    return ", ".join(parts)


async def export_orders_to_excel(
    db: AsyncSession,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[str] = None,
    service_type: Optional[str] = None
) -> bytes:
    """
    Export orders to Excel format with optional filters.
    Filters by order creation date (sale date), not by flight departure date.
    """
    # Build query with eager loading
    query = select(Order).options(
        selectinload(Order.services).selectinload(Service.origin_location),
        selectinload(Order.services).selectinload(Service.destination_location),
        selectinload(Order.customer),
        selectinload(Order.user)
    )

    # Apply filters to Order created_at (sale date), not flight departure date
    if start_date:
        start_datetime = datetime.combine(start_date, datetime.min.time())
        query = query.where(Order.created_at >= start_datetime)
    if end_date:
        end_datetime = datetime.combine(end_date, datetime.max.time())
        query = query.where(Order.created_at <= end_datetime)

    # Execute query first to get orders
    result = await db.execute(query)
    orders = result.scalars().unique().all()

    # Filter services within orders if needed
    if status or service_type:
        filtered_orders = []
        for order in orders:
            filtered_services = order.services
            if status:
                filtered_services = [s for s in filtered_services if s.status == status.lower()]
            if service_type:
                filtered_services = [s for s in filtered_services if s.service_type.value == service_type.upper()]

            if filtered_services:
                # Create a copy with filtered services
                order.services = filtered_services
                filtered_orders.append(order)
        orders = filtered_orders

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders Export"

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Order Number", "Customer Name", "Document", "Email", "Phone",
        "Service Type", "Service Name", "Departure", "Arrival", "Origin", "Destination",
        "Status", "Cost Price", "Sale Price", "Profit", "Created At"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_style

    # Data rows
    row_num = 2
    for order in orders:
        for service in order.services:
            ws.cell(row=row_num, column=1, value=order.order_number)
            ws.cell(row=row_num, column=2, value=order.customer.full_name)
            ws.cell(row=row_num, column=3, value=order.customer.document_id or "")
            ws.cell(row=row_num, column=4, value=order.customer.email or "")
            ws.cell(row=row_num, column=5, value=order.customer.phone_number or "")
            ws.cell(row=row_num, column=6, value=service.service_type.value if hasattr(service.service_type, 'value') else str(service.service_type))
            ws.cell(row=row_num, column=7, value=service.name)
            ws.cell(row=row_num, column=8, value=service.departure_datetime.strftime("%Y-%m-%d %H:%M") if service.departure_datetime else "")
            ws.cell(row=row_num, column=9, value=service.arrival_datetime.strftime("%Y-%m-%d %H:%M") if service.arrival_datetime else "")
            ws.cell(row=row_num, column=10, value=format_location(service.origin_location))
            ws.cell(row=row_num, column=11, value=format_location(service.destination_location))
            ws.cell(row=row_num, column=12, value=service.status)
            ws.cell(row=row_num, column=13, value=float(order.total_cost_price))
            ws.cell(row=row_num, column=14, value=float(order.total_sale_price))
            ws.cell(row=row_num, column=15, value=float(order.total_profit))
            ws.cell(row=row_num, column=16, value=order.created_at.strftime("%Y-%m-%d %H:%M"))

            # Apply borders
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).border = border_style

            row_num += 1

    # Adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file.getvalue()


async def export_orders_to_pdf(
    db: AsyncSession,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[str] = None,
    service_type: Optional[str] = None
) -> bytes:
    """
    Export orders to PDF format with optional filters.
    Filters by order creation date (sale date), not by flight departure date.
    """
    # Build query with eager loading
    query = select(Order).options(
        selectinload(Order.services).selectinload(Service.origin_location),
        selectinload(Order.services).selectinload(Service.destination_location),
        selectinload(Order.customer),
        selectinload(Order.user)
    )

    # Apply filters to Order created_at (sale date), not flight departure date
    if start_date:
        start_datetime = datetime.combine(start_date, datetime.min.time())
        query = query.where(Order.created_at >= start_datetime)
    if end_date:
        end_datetime = datetime.combine(end_date, datetime.max.time())
        query = query.where(Order.created_at <= end_datetime)

    # Execute query first to get orders
    result = await db.execute(query)
    orders = result.scalars().unique().all()

    # Filter services within orders if needed
    if status or service_type:
        filtered_orders = []
        for order in orders:
            filtered_services = order.services
            if status:
                filtered_services = [s for s in filtered_services if s.status == status.lower()]
            if service_type:
                filtered_services = [s for s in filtered_services if s.service_type.value == service_type.upper()]

            if filtered_services:
                # Create a copy with filtered services
                order.services = filtered_services
                filtered_orders.append(order)
        orders = filtered_orders

    # Create PDF
    pdf_file = io.BytesIO()
    doc = SimpleDocTemplate(pdf_file, pagesize=A4)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4472C4'),
        alignment=TA_CENTER,
        spaceAfter=30
    )

    # Title
    elements.append(Paragraph("Travel Sales Report", title_style))
    elements.append(Spacer(1, 0.3*inch))

    # Summary section
    summary_data = [
        ['Report Information', ''],
        ['Generated:', datetime.now().strftime("%Y-%m-%d %H:%M")],
        ['Total Orders:', str(len(orders))],
    ]

    if start_date:
        summary_data.append(['Start Date:', start_date.strftime("%Y-%m-%d")])
    if end_date:
        summary_data.append(['End Date:', end_date.strftime("%Y-%m-%d")])
    if status:
        summary_data.append(['Status Filter:', status])
    if service_type:
        summary_data.append(['Service Type:', service_type])

    summary_table = Table(summary_data, colWidths=[2*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 0.5*inch))

    # Orders table
    if orders:
        # Table header
        data = [['Order #', 'Customer', 'Service', 'Date', 'Status', 'Cost', 'Sale', 'Profit']]

        # Table rows
        for order in orders:
            for service in order.services:
                service_type = service.service_type.value if hasattr(service.service_type, 'value') else str(service.service_type)
                data.append([
                    str(order.order_number)[:10],
                    order.customer.full_name[:20],
                    service_type,
                    service.departure_datetime.strftime("%Y-%m-%d") if service.departure_datetime else "",
                    service.status,
                    f"${float(order.total_cost_price):.2f}",
                    f"${float(order.total_sale_price):.2f}",
                    f"${float(order.total_profit):.2f}"
                ])

        # Create table
        orders_table = Table(data, colWidths=[0.8*inch, 1.3*inch, 0.8*inch, 0.9*inch, 0.8*inch, 0.7*inch, 0.7*inch, 0.7*inch])
        orders_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))

        elements.append(Paragraph("Orders Details", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        elements.append(orders_table)

        # Calculate totals
        total_cost = sum(float(order.total_cost_price) for order in orders)
        total_sale = sum(float(order.total_sale_price) for order in orders)
        total_profit = sum(float(order.total_profit) for order in orders)

        elements.append(Spacer(1, 0.3*inch))

        totals_data = [
            ['Total Cost Price:', f"${total_cost:.2f}"],
            ['Total Sale Price:', f"${total_sale:.2f}"],
            ['Total Profit:', f"${total_profit:.2f}"]
        ]

        totals_table = Table(totals_data, colWidths=[2*inch, 2*inch])
        totals_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E7E6E6')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(totals_table)
    else:
        elements.append(Paragraph("No orders found matching the criteria.", styles['Normal']))

    # Build PDF
    doc.build(elements)
    pdf_file.seek(0)

    return pdf_file.getvalue()
